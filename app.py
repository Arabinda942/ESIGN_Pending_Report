"""
ESIGN Report Generator
=======================
Two distinct report types:

  TYPE 1 — Pending Report  (Main.xlsx + L1L2L3BranchCode.xlsx)
    The original "ESIGN Pending Report": filters down to records that are
    still stuck (status == 'W', EsigncompletedDate empty, Pan not empty),
    breaks each one into a pending sub-stage (Address/Profile/Bank/
    Exchange/IPV/PDF/Esign), and looks up RM Mobile / Name / L1 / L2 /
    Br code via the hierarchy file. This is the report your team already
    uses — logic is unchanged from before.

  TYPE 2 — Raw File Report  (Main.xlsx only)
    Takes the raw eKYC export as-is, with NO row filtering, and buckets
    every record into Completed / Rejected / Pending based on the
    'status' column (A = Completed, R = Rejected, W = Pending). Pending
    records additionally get the same sub-stage breakdown as Type 1.
    No RM hierarchy is used or needed here — this is a full-file status
    snapshot, not a "who's it stuck with" report.

Both types get the same UX: generate -> results screen -> Download Excel
or View Interactive Dashboard.

Run:
    pip install flask pandas openpyxl --break-system-packages
    python app.py
Then open http://127.0.0.1:5000
"""

import io
import re
import json
import uuid
import calendar
from collections import Counter
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from flask import (
    Flask, request, send_file, render_template_string, redirect, url_for,
)

app = Flask(__name__)

# ----------------------------------------------------------------------
# Column configuration
# ----------------------------------------------------------------------

MAIN_COLUMNS_KEPT = [
    "EkycID", "Pan", "DOB", "Mobile", "Creationdate", "Name", "Address",
    "Profile", "Bank", "Exchange", "IPV", "PDF", "DigiSign",
    "RequestStatusPath", "AccountMode", "SourceEmpName", "gender",
    "status", "EsigncompletedDate",
]

# --- Type 1 (Pending Report) output shape ---
PENDING_DERIVED_COLUMNS = ["Pending At", "RM Mobile ", "Name", "L1", "L2", "Br code"]
PENDING_OUTPUT_HEADERS = MAIN_COLUMNS_KEPT + PENDING_DERIVED_COLUMNS

PENDING_COLUMN_WIDTHS = {  # by column letter, captured from the original Output.xlsx
    "A": 6.71, "B": 13.43, "C": 10.43, "D": 11.0, "E": 15.57, "F": 38.29,
    "G": 8.14, "H": 7.0, "I": 5.29, "J": 9.29, "K": 4.0, "L": 4.43,
    "M": 8.14, "N": 48.29, "O": 20.57, "P": 47.14, "Q": 7.29, "R": 6.29,
    "S": 19.57, "T": 32.0, "U": 13.57, "V": 31.29, "W": 30.86, "X": 29.43,
    "Y": 11.14,
}

# --- Type 2 (Raw File Report) output shape ---
RAW_DERIVED_COLUMNS = ["Overall Status", "Pending At"]
RAW_OUTPUT_HEADERS = MAIN_COLUMNS_KEPT + RAW_DERIVED_COLUMNS

HEADER_WIDTHS = {  # by header name, used for the Raw File Report
    "EkycID": 6.71, "Pan": 13.43, "DOB": 10.43, "Mobile": 11.0,
    "Creationdate": 15.57, "Name": 38.29, "Address": 8.14, "Profile": 7.0,
    "Bank": 5.29, "Exchange": 9.29, "IPV": 4.0, "PDF": 4.43,
    "DigiSign": 8.14, "RequestStatusPath": 48.29, "AccountMode": 20.57,
    "SourceEmpName": 47.14, "gender": 7.29, "status": 6.29,
    "EsigncompletedDate": 19.57, "Overall Status": 16.0, "Pending At": 32.0,
}
DEFAULT_COLUMN_WIDTH = 14.0

DOB_FORMAT = "[$-14009]dd/mm/yyyy;@"
DATETIME_FORMAT = "m/d/yy h:mm"

MOBILE_IN_PARENS = re.compile(r"\((\d{7,15})\)\s*$")

STAGE_ORDER = [
    "Address Stage", "Profile Stage", "Bank Stage", "Exchange Stage",
    "Selfie/IPV not done", "IPV Approved PDF not done",
    "IPV done PDF not done", "All stages done but Esign not done",
]

STATUS_CODE_MAP = {"A": "Completed", "R": "Rejected", "W": "Pending"}
STATUS_ORDER = ["Completed", "Rejected", "Pending", "Unknown"]

# In-memory store for generated reports, keyed by a short random id.
# Local single-user tool, so no expiry/cleanup is implemented.
REPORTS = {}


# ----------------------------------------------------------------------
# Shared helpers
# ----------------------------------------------------------------------

def _blank(v):
    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    if str(v).strip() == "":
        return True
    return False


def load_main(file_stream):
    return pd.read_excel(file_stream, sheet_name=0)


def select_columns(df):
    """Pick MAIN_COLUMNS_KEPT and normalize Mobile to text (avoids
    scientific notation / leading-zero loss on re-export)."""
    out = df[MAIN_COLUMNS_KEPT].copy()
    out["Mobile"] = out["Mobile"].apply(
        lambda v: str(int(v)) if pd.notna(v) and str(v).strip() != "" and float(str(v)).is_integer()
        else ("" if pd.isna(v) else str(v))
    )
    return out


def parse_dates_robust(series):
    """Parse a date column that may already be proper datetimes, text
    dates, or raw Excel serial numbers (happens when the source cells
    are formatted as 'General' instead of a date format, so pandas
    reads them as plain floats). Handles all three cases."""
    dt = pd.to_datetime(series, errors="coerce")
    numeric = pd.to_numeric(series, errors="coerce")
    looks_numeric = numeric.notna().mean() > 0.5
    parsed_well = dt.notna().mean() > 0.5 and dt.dt.year.between(1990, 2100).mean() > 0.5
    if looks_numeric and not parsed_well:
        dt = pd.to_datetime(numeric, unit="D", origin="1899-12-30", errors="coerce")
    return dt


def ordinal(n):
    if 11 <= (n % 100) <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n:02d}{suffix}"


def build_title(main_df, label):
    dates = parse_dates_robust(main_df["Creationdate"]).dropna()
    counts = Counter(zip(dates.dt.year, dates.dt.month))
    (year, month), _ = counts.most_common(1)[0]

    in_month = dates[(dates.dt.year == year) & (dates.dt.month == month)]
    first_day = int(in_month.dt.day.min())
    last_day = int(in_month.dt.day.max())
    month_name = calendar.month_name[month]
    return (
        f"{label} from {ordinal(first_day)} {month_name} {year} "
        f"to {ordinal(last_day)} {month_name} {year}"
    )


def pending_stage(row):
    """The sub-stage a still-pending record is stuck at."""
    if _blank(row["Address"]):
        return "Address Stage"
    if _blank(row["Profile"]):
        return "Profile Stage"
    if _blank(row["Bank"]):
        return "Bank Stage"
    if _blank(row["Exchange"]):
        return "Exchange Stage"
    ipv = "" if _blank(row["IPV"]) else str(row["IPV"]).strip().upper()
    if ipv == "":
        return "Selfie/IPV not done"
    if _blank(row["PDF"]):
        if ipv == "A":
            return "IPV Approved PDF not done"
        return "IPV done PDF not done"
    return "All stages done but Esign not done"


def parse_emp_display_name(source_emp_name):
    """Strip the trailing '(mobile number)' off SourceEmpName, e.g.
    'AVRADEEP SANYAL(9830383669)' -> 'AVRADEEP SANYAL'."""
    if _blank(source_emp_name):
        return "Unknown"
    s = str(source_emp_name).strip()
    m = MOBILE_IN_PARENS.search(s)
    if m:
        s = s[: m.start()].strip()
    return s or "Unknown"


def _top_n(series, n, other_label):
    cleaned = series.apply(lambda v: other_label if _blank(v) else v)
    vc = cleaned.value_counts().head(n)
    return [str(x) for x in vc.index], [int(x) for x in vc.values]


def _digits(v):
    if _blank(v):
        return None
    s = re.sub(r"\D", "", str(v))
    return s or None


def write_excel(df, title, headers, name_key_overrides=None, widths_by_letter=None, widths_by_name=None):
    """Generic Excel writer shared by both report types.
    name_key_overrides: dict header -> alternate dataframe column name
      (needed because the Pending Report has two columns literally
      called 'Name' — the client's Name and the RM's Name)."""
    name_key_overrides = name_key_overrides or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    bold12 = Font(name="Calibri", size=12, bold=True)
    bold11 = Font(name="Calibri", size=11, bold=True)

    ws.cell(row=1, column=1, value=title).font = bold12
    for c, header in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=header).font = bold11

    for r, (_, row) in enumerate(df.iterrows(), start=4):
        for c, header in enumerate(headers, start=1):
            col_key = name_key_overrides.get((header, c), header)
            value = row.get(col_key)
            if pd.isna(value):
                value = ""
            cell = ws.cell(row=r, column=c, value=value)
            if header == "DOB":
                cell.number_format = DOB_FORMAT
            elif header == "Creationdate":
                cell.number_format = DATETIME_FORMAT

    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        if widths_by_letter and letter in widths_by_letter:
            ws.column_dimensions[letter].width = widths_by_letter[letter]
        elif widths_by_name:
            ws.column_dimensions[letter].width = widths_by_name.get(header, DEFAULT_COLUMN_WIDTH)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ----------------------------------------------------------------------
# TYPE 1 — Pending Report (Main.xlsx + L1L2L3BranchCode.xlsx)
# ----------------------------------------------------------------------

def filter_pending_rows(df):
    mask = (
        (df["status"].astype(str).str.strip() == "W")
        & (df["EsigncompletedDate"].isna())
        & (df["Pan"].notna())
    )
    return select_columns(df.loc[mask])


def extract_mobile(source_emp_name):
    if _blank(source_emp_name):
        return None
    m = MOBILE_IN_PARENS.search(str(source_emp_name).strip())
    return m.group(1) if m else None


def load_rm_lookup(file_stream):
    """Returns dict: mobile-digits -> {Name, L1, L2, Br code}"""
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    ws = wb.active

    header = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is not None:
            header[str(val).strip()] = col

    col_mobile1 = 1
    col_mobile2 = 2
    col_name = header.get("Name")
    col_l1 = header.get("L1")
    col_l2 = header.get("L2")
    col_br = header.get("Br code")
    col_mobileno_named = header.get("Mobileno", col_mobile1)

    lookup = {}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=col_name).value if col_name else None
        if _blank(name):
            continue
        info = {
            "Name": name,
            "L1": ws.cell(row=row, column=col_l1).value if col_l1 else None,
            "L2": ws.cell(row=row, column=col_l2).value if col_l2 else None,
            "Br code": ws.cell(row=row, column=col_br).value if col_br else None,
        }
        for col_idx in (col_mobile1, col_mobile2, col_mobileno_named):
            mob = _digits(ws.cell(row=row, column=col_idx).value)
            if mob and mob not in lookup:
                lookup[mob] = info
    return lookup


def apply_rm_lookup(df, rm_lookup):
    rm_mobiles, names, l1s, l2s, brs = [], [], [], [], []
    for _, row in df.iterrows():
        mob = extract_mobile(row["SourceEmpName"])
        info = rm_lookup.get(mob) if mob else None
        if info:
            rm_mobiles.append(int(mob) if mob.isdigit() else mob)
            names.append(info["Name"])
            l1s.append(info["L1"])
            l2s.append(info["L2"])
            brs.append(info["Br code"])
        else:
            rm_mobiles.append(None)
            names.append(None)
            l1s.append(None)
            l2s.append(None)
            brs.append(None)
    df["RM Mobile "] = rm_mobiles
    df["Name.1"] = names   # renamed to "Name" at write time
    df["L1"] = l1s
    df["L2"] = l2s
    df["Br code"] = brs
    return df


def is_direct_sourced(source_emp_name):
    """True when Main.xlsx itself already says this record has no RM —
    SourceEmpName is literally the word 'Direct' (no name / mobile at
    all), as opposed to a named employee whose mobile just didn't match
    a row in L1L2L3BranchCode.xlsx (a genuine 'No RM Match' data gap)."""
    return parse_emp_display_name(source_emp_name) == "Direct"


def _attribution_bucket(value_series, direct_mask, other_label="No RM Match"):
    """Fill blanks (no RM lookup match) with 'Direct' where Main.xlsx's
    own SourceEmpName says so, and with other_label otherwise — so a
    genuine lookup miss is never silently mislabeled as Direct."""
    out = []
    for value, is_direct in zip(value_series, direct_mask):
        if _blank(value):
            out.append("Direct" if is_direct else other_label)
        else:
            out.append(value)
    return pd.Series(out, index=value_series.index)


def build_pending_summary(df, title):
    total = len(df)

    stage_counts = df["Pending At"].value_counts()
    stage_labels = [s for s in STAGE_ORDER if s in stage_counts.index]
    stage_labels += [s for s in stage_counts.index if s not in STAGE_ORDER]
    stage_values = [int(stage_counts.get(s, 0)) for s in stage_labels]

    direct_mask = df["SourceEmpName"].apply(is_direct_sourced)
    l1_labels, l1_values = _top_n(_attribution_bucket(df["L1"], direct_mask), 12, "No RM Match")
    l2_labels, l2_values = _top_n(_attribution_bucket(df["L2"], direct_mask), 12, "No RM Match")
    branch_labels, branch_values = _top_n(_attribution_bucket(df["Br code"], direct_mask), 15, "No RM Match")
    rm_labels, rm_values = _top_n(_attribution_bucket(df["Name.1"], direct_mask), 15, "No RM Match")
    gender_labels, gender_values = _top_n(df["gender"], 6, "Unspecified")
    mode_labels, mode_values = _top_n(df["AccountMode"], 8, "Unspecified")

    dates = parse_dates_robust(df["Creationdate"]).dt.date
    trend = dates.value_counts().dropna().sort_index()
    trend_labels = [d.strftime("%d %b") for d in trend.index]
    trend_values = [int(x) for x in trend.values]

    unmatched_rm = int(df["Name.1"].apply(_blank).sum())
    esign_only = int((df["Pending At"] == "All stages done but Esign not done").sum())

    return {
        "report_type": "pending",
        "title": title,
        "total": total,
        "unmatched_rm": unmatched_rm,
        "esign_only": esign_only,
        "generated_at": datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "stage": {"labels": stage_labels, "values": stage_values},
        "l1": {"labels": l1_labels, "values": l1_values},
        "l2": {"labels": l2_labels, "values": l2_values},
        "branch": {"labels": branch_labels, "values": branch_values},
        "rm": {"labels": rm_labels, "values": rm_values},
        "gender": {"labels": gender_labels, "values": gender_values},
        "mode": {"labels": mode_labels, "values": mode_values},
        "trend": {"labels": trend_labels, "values": trend_values},
    }


def generate_pending_report(main_stream, l1l2l3_stream):
    main_df = load_main(main_stream)
    pending_df = filter_pending_rows(main_df)
    pending_df["Pending At"] = pending_df.apply(pending_stage, axis=1)

    rm_lookup = load_rm_lookup(l1l2l3_stream)
    pending_df = apply_rm_lookup(pending_df, rm_lookup)

    title = build_title(main_df, "ESIGN Pending Report")
    summary = build_pending_summary(pending_df, title)

    name_overrides = {("Name", c): "Name.1" for c, h in enumerate(PENDING_OUTPUT_HEADERS, start=1) if h == "Name" and c > 6}
    output_buf = write_excel(
        pending_df, title, PENDING_OUTPUT_HEADERS,
        name_key_overrides=name_overrides, widths_by_letter=PENDING_COLUMN_WIDTHS,
    )
    return output_buf, len(pending_df), summary


# ----------------------------------------------------------------------
# TYPE 2 — Raw File Report (Main.xlsx only; Pending/Rejected/Completed)
# ----------------------------------------------------------------------

def overall_status(row):
    """Completed / Rejected are decided by status code alone. 'Pending'
    additionally requires EsigncompletedDate to be blank AND Pan to be
    filled in — the exact same test the Pending Report (Type 1) uses.
    A status='W' row that hasn't reached PAN yet, or that oddly already
    has an EsigncompletedDate, is not a genuine actionable pending case
    and is bucketed as Unknown instead, so the two report types agree
    on how many records are truly pending."""
    s = "" if _blank(row["status"]) else str(row["status"]).strip().upper()
    if s == "A":
        return "Completed"
    if s == "R":
        return "Rejected"
    if s == "W":
        if _blank(row["EsigncompletedDate"]) and not _blank(row["Pan"]):
            return "Pending"
        return "Unknown"
    return "Unknown"


def build_raw_summary(df, title):
    total = len(df)
    status_counts = df["Overall Status"].value_counts()
    status_labels = [s for s in STATUS_ORDER if s in status_counts.index]
    status_values = [int(status_counts.get(s, 0)) for s in status_labels]

    completed_count = int(status_counts.get("Completed", 0))
    rejected_count = int(status_counts.get("Rejected", 0))
    pending_count = int(status_counts.get("Pending", 0))

    pending_df = df[df["Overall Status"] == "Pending"]
    stage_counts = pending_df["Pending At"].value_counts()
    stage_labels = [s for s in STAGE_ORDER if s in stage_counts.index]
    stage_labels += [s for s in stage_counts.index if s not in STAGE_ORDER]
    stage_values = [int(stage_counts.get(s, 0)) for s in stage_labels]
    esign_only = int((pending_df["Pending At"] == "All stages done but Esign not done").sum())

    rejected_df = df[df["Overall Status"] == "Rejected"]
    reject_labels, reject_values = _top_n(rejected_df["RequestStatusPath"], 8, "Unspecified")

    gender_labels, gender_values = _top_n(df["gender"], 6, "Unspecified")
    mode_labels, mode_values = _top_n(df["AccountMode"], 8, "Unspecified")

    emp_names = df["SourceEmpName"].apply(parse_emp_display_name)
    source_emp_labels, source_emp_values = _top_n(emp_names, 10, "Unknown")

    # RM Wise (raw report has no RM hierarchy file, but SourceEmpName IS
    # the sourcing RM's name — or literally 'Direct' when there is none —
    # so top RMs, split by status, is available straight from Main.xlsx.
    rm_top_labels = emp_names.value_counts().head(12).index.tolist()
    rm_series = {s: [] for s in ["Completed", "Rejected", "Pending"]}
    for lbl in rm_top_labels:
        row_mask = emp_names == lbl
        for s in rm_series:
            rm_series[s].append(int((row_mask & (df["Overall Status"] == s)).sum()))

    # Daily trend, split by status (3 aligned series over the same date axis)
    dates = parse_dates_robust(df["Creationdate"]).dt.date
    trend_df = pd.DataFrame({"date": dates, "status": df["Overall Status"]}).dropna(subset=["date"])
    pivot = trend_df.pivot_table(index="date", columns="status", values="status", aggfunc="count", fill_value=0)
    pivot = pivot.sort_index()
    trend_labels = [d.strftime("%d %b") for d in pivot.index]
    trend_series = {
        s: [int(x) for x in pivot[s]] if s in pivot.columns else [0] * len(pivot)
        for s in ["Completed", "Rejected", "Pending"]
    }

    # Average turnaround (Completed records only, where both dates parse cleanly)
    completed_df = df[df["Overall Status"] == "Completed"]
    created = parse_dates_robust(completed_df["Creationdate"])
    finished = parse_dates_robust(completed_df["EsigncompletedDate"])
    delta_days = (finished - created).dt.total_seconds() / 86400.0
    delta_days = delta_days.dropna()
    delta_days = delta_days[(delta_days >= 0) & (delta_days < 365)]
    avg_turnaround = round(float(delta_days.mean()), 2) if len(delta_days) else None

    return {
        "report_type": "raw",
        "title": title,
        "total": total,
        "completed_count": completed_count,
        "rejected_count": rejected_count,
        "pending_count": pending_count,
        "esign_only": esign_only,
        "avg_turnaround": avg_turnaround,
        "generated_at": datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "status": {"labels": status_labels, "values": status_values},
        "stage": {"labels": stage_labels, "values": stage_values},
        "reject_reasons": {"labels": reject_labels, "values": reject_values},
        "gender": {"labels": gender_labels, "values": gender_values},
        "mode": {"labels": mode_labels, "values": mode_values},
        "source_emp": {"labels": source_emp_labels, "values": source_emp_values},
        "rm": {"labels": rm_top_labels, "series": rm_series},
        "trend": {"labels": trend_labels, "series": trend_series},
    }


def generate_raw_report(main_stream):
    main_df = load_main(main_stream)
    raw_df = select_columns(main_df)
    raw_df["Overall Status"] = raw_df.apply(overall_status, axis=1)
    raw_df["Pending At"] = raw_df.apply(
        lambda r: pending_stage(r) if r["Overall Status"] == "Pending" else "", axis=1
    )

    title = build_title(main_df, "ESIGN Status Report (Pending / Rejected / Completed)")
    summary = build_raw_summary(raw_df, title)

    output_buf = write_excel(
        raw_df, title, RAW_OUTPUT_HEADERS, widths_by_name=HEADER_WIDTHS,
    )
    return output_buf, len(raw_df), summary


# ----------------------------------------------------------------------
# Shared styling (green / black theme, with status accent colors)
# ----------------------------------------------------------------------

BASE_STYLE = """
:root {
  --bg: #0b0f14;
  --panel: #121822;
  --panel-2: #0d1319;
  --panel-border: #1f2937;
  --accent: #3ddc97;
  --accent-dim: #22916b;
  --accent-soft: rgba(61,220,151,0.12);
  --text: #e6edf3;
  --muted: #8b98a5;
  --danger: #ff6b6b;
  --warn: #f4c95d;
  --info: #5fb3ff;
  --info-soft: rgba(95,179,255,0.1);
}
* { box-sizing: border-box; }
body {
  margin: 0;
  min-height: 100vh;
  background: radial-gradient(1200px 600px at 50% -10%, #16202c 0%, var(--bg) 60%);
  color: var(--text);
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, sans-serif;
}
a { color: inherit; }
.note { font-size: 0.78rem; color: var(--muted); line-height: 1.5; }
.error {
  background: rgba(255,107,107,0.1);
  border: 1px solid rgba(255,107,107,0.35);
  color: var(--danger);
  padding: 10px 14px;
  border-radius: 8px;
  font-size: 0.85rem;
}
.info-banner {
  background: var(--info-soft);
  border: 1px solid rgba(95,179,255,0.35);
  color: var(--info);
  padding: 10px 14px;
  border-radius: 8px;
  font-size: 0.82rem;
  line-height: 1.5;
}
"""


# ----------------------------------------------------------------------
# Page 1: Upload form (Pending Report tab / Raw File tab)
# ----------------------------------------------------------------------

PAGE = """
<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>ESIGN Report Generator</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
""" + BASE_STYLE + """
  body { display: flex; align-items: center; justify-content: center; padding: 24px; }
  .card {
    width: 100%;
    max-width: 580px;
    background: var(--panel);
    border: 1px solid var(--panel-border);
    border-radius: 14px;
    padding: 32px;
    box-shadow: 0 20px 60px rgba(0,0,0,0.45);
  }
  h1 { font-size: 1.35rem; margin: 0 0 4px 0; letter-spacing: -0.02em; }
  .sub { color: var(--muted); font-size: 0.9rem; margin-bottom: 22px; }

  .tabs {
    display: grid; grid-template-columns: 1fr 1fr; gap: 8px;
    background: var(--panel-2); border: 1px solid var(--panel-border);
    border-radius: 11px; padding: 5px; margin-bottom: 24px;
  }
  .tab {
    text-align: center; padding: 10px 8px; border-radius: 8px;
    font-size: 0.82rem; font-weight: 700; cursor: pointer; color: var(--muted);
    transition: background .15s ease, color .15s ease; user-select: none;
  }
  .tab span { font-weight: 500; opacity: .75; }
  .tab.active { background: var(--accent); color: #04140d; }

  label {
    display: block; font-size: 0.82rem; color: var(--muted);
    margin: 18px 0 6px; text-transform: uppercase; letter-spacing: 0.04em;
  }
  .dropzone {
    border: 1.5px dashed var(--panel-border);
    border-radius: 10px; padding: 16px; background: var(--panel-2);
    transition: border-color .15s ease, background .15s ease;
  }
  .dropzone:hover { border-color: var(--accent-dim); }
  input[type=file] { width: 100%; color: var(--text); font-size: 0.88rem; }
  input[type=file]::file-selector-button {
    background: var(--accent-dim); color: #05130d; border: none;
    padding: 8px 14px; border-radius: 7px; font-weight: 600;
    margin-right: 12px; cursor: pointer;
  }
  button {
    width: 100%; margin-top: 28px; padding: 13px; border: none;
    border-radius: 10px; background: var(--accent); color: #04140d;
    font-weight: 700; font-size: 0.95rem; letter-spacing: 0.01em;
    cursor: pointer; transition: transform .08s ease, filter .15s ease;
  }
  button:hover { filter: brightness(1.08); }
  button:active { transform: scale(0.99); }
  .error, .note, .info-banner { margin-top: 18px; }
  #panel-pending, #panel-raw { display: none; }
  #panel-pending.show, #panel-raw.show { display: block; }
  .legend-row { display: flex; gap: 14px; margin-top: 10px; flex-wrap: wrap; }
  .legend-chip { display: flex; align-items: center; gap: 6px; font-size: 0.74rem; color: var(--muted); }
  .legend-dot { width: 8px; height: 8px; border-radius: 50%; }
</style>
</head>
<body>
  <div class="card">
    <h1>ESIGN Report Generator</h1>
    <div class="sub">Generate an Excel report and see it as an instant dashboard.</div>
    {% if error %}<div class="error">{{ error }}</div>{% endif %}

    <div class="tabs">
      <div class="tab active" id="tab-pending" onclick="setMode('pending')">Pending Report<br><span>Main + RM Hierarchy</span></div>
      <div class="tab" id="tab-raw" onclick="setMode('raw')">Raw File<br><span>Main.xlsx only — all statuses</span></div>
    </div>

    <form method="post" enctype="multipart/form-data" id="uploadForm">
      <input type="hidden" name="mode" id="modeField" value="pending">

      <div id="panel-pending" class="show">
        <label>Main.xlsx (raw eKYC export)</label>
        <div class="dropzone"><input type="file" name="main_file" id="mainFilePending" accept=".xlsx"></div>

        <label>L1L2L3BranchCode.xlsx (RM hierarchy)</label>
        <div class="dropzone"><input type="file" name="l1l2l3_file" id="rmFile" accept=".xlsx"></div>

        <div class="note">
          Filters to records still pending: status = "W", EsigncompletedDate empty, Pan not empty.<br>
          Output: EkycID … EsigncompletedDate + Pending At, RM Mobile, Name, L1, L2, Br code.
        </div>
      </div>

      <div id="panel-raw">
        <label>Main.xlsx (raw eKYC export, no filtering)</label>
        <div class="dropzone"><input type="file" name="main_file_raw" id="mainFileRaw" accept=".xlsx"></div>

        <div class="info-banner">
          Every record in the file is kept and bucketed by status: <b>Completed</b> (A), <b>Rejected</b> (R),
          <b>Pending</b> (W). Pending records also get the same Address/Profile/Bank/Exchange/IPV/PDF/Esign
          sub-stage breakdown. No RM hierarchy file needed.
        </div>
        <div class="legend-row">
          <span class="legend-chip"><span class="legend-dot" style="background:#3ddc97;"></span>Completed</span>
          <span class="legend-chip"><span class="legend-dot" style="background:#ff6b6b;"></span>Rejected</span>
          <span class="legend-chip"><span class="legend-dot" style="background:#f4c95d;"></span>Pending</span>
        </div>
      </div>

      <button type="submit">Generate Report</button>
    </form>
  </div>

  <script>
    function setMode(mode) {
      const pending = document.getElementById('panel-pending');
      const raw = document.getElementById('panel-raw');
      const tabPending = document.getElementById('tab-pending');
      const tabRaw = document.getElementById('tab-raw');
      const modeField = document.getElementById('modeField');
      const mainFilePending = document.getElementById('mainFilePending');
      const rmFile = document.getElementById('rmFile');
      const mainFileRaw = document.getElementById('mainFileRaw');

      modeField.value = mode;
      if (mode === 'pending') {
        pending.classList.add('show'); raw.classList.remove('show');
        tabPending.classList.add('active'); tabRaw.classList.remove('active');
        mainFilePending.required = true; rmFile.required = true; mainFileRaw.required = false;
      } else {
        pending.classList.remove('show'); raw.classList.add('show');
        tabPending.classList.remove('active'); tabRaw.classList.add('active');
        mainFilePending.required = false; rmFile.required = false; mainFileRaw.required = true;
      }
    }
    setMode('pending');
  </script>
</body>
</html>
"""


# ----------------------------------------------------------------------
# Page 2: Results (choose Download or Dashboard)
# ----------------------------------------------------------------------

RESULTS_PAGE = """
<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Report Ready</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
""" + BASE_STYLE + """
  body { display: flex; align-items: center; justify-content: center; padding: 24px; }
  .card {
    width: 100%; max-width: 620px; background: var(--panel);
    border: 1px solid var(--panel-border); border-radius: 14px;
    padding: 36px; box-shadow: 0 20px 60px rgba(0,0,0,0.45);
  }
  .badge {
    display: inline-flex; align-items: center; gap: 8px;
    background: var(--accent-soft); border: 1px solid rgba(61,220,151,0.35);
    color: var(--accent); padding: 6px 12px; border-radius: 999px;
    font-size: 0.78rem; font-weight: 600; letter-spacing: 0.02em;
  }
  .dot { width: 7px; height: 7px; border-radius: 50%; background: var(--accent); }
  h1 { font-size: 1.25rem; margin: 18px 0 4px 0; letter-spacing: -0.02em; }
  .sub { color: var(--muted); font-size: 0.88rem; margin-bottom: 16px; }
  .stats { display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin-bottom: 20px; }
  .stat {
    background: var(--panel-2); border: 1px solid var(--panel-border);
    border-radius: 10px; padding: 14px 12px; text-align: center;
  }
  .stat .num { font-size: 1.5rem; font-weight: 800; letter-spacing: -0.02em; }
  .stat .num.green { color: var(--accent); }
  .stat .num.red { color: var(--danger); }
  .stat .num.amber { color: var(--warn); }
  .stat .lbl { font-size: 0.7rem; color: var(--muted); margin-top: 2px; text-transform: uppercase; letter-spacing: 0.03em; }
  .actions { display: grid; gap: 12px; margin-top: 10px; }
  a.btn {
    display: flex; align-items: center; justify-content: center; gap: 8px;
    padding: 15px; border-radius: 10px; font-weight: 700; font-size: 0.95rem;
    text-decoration: none; letter-spacing: 0.01em; transition: transform .08s ease, filter .15s ease;
  }
  a.btn:active { transform: scale(0.99); }
  .btn-primary { background: var(--accent); color: #04140d; }
  .btn-primary:hover { filter: brightness(1.08); }
  .btn-outline { background: transparent; color: var(--accent); border: 1.5px solid var(--accent-dim); }
  .btn-outline:hover { background: var(--accent-soft); }
  .back { display: block; text-align: center; margin-top: 22px; color: var(--muted); font-size: 0.82rem; text-decoration: none; }
  .back:hover { color: var(--text); }
</style>
</head>
<body>
  <div class="card">
    {% if summary.report_type == "raw" %}
      <span class="badge"><span class="dot"></span>Raw file report generated</span>
      <h1>{{ summary.title }}</h1>
      <div class="sub">{{ row_count }} record(s) processed · generated {{ summary.generated_at }}</div>
      <div class="stats">
        <div class="stat"><div class="num green">{{ summary.completed_count }}</div><div class="lbl">Completed</div></div>
        <div class="stat"><div class="num red">{{ summary.rejected_count }}</div><div class="lbl">Rejected</div></div>
        <div class="stat"><div class="num amber">{{ summary.pending_count }}</div><div class="lbl">Pending</div></div>
      </div>
    {% else %}
      <span class="badge"><span class="dot"></span>Pending report generated</span>
      <h1>{{ summary.title }}</h1>
      <div class="sub">{{ row_count }} pending record(s) found · generated {{ summary.generated_at }}</div>
      <div class="stats">
        <div class="stat"><div class="num green">{{ row_count }}</div><div class="lbl">Total Pending</div></div>
        <div class="stat"><div class="num green">{{ summary.esign_only }}</div><div class="lbl">Esign Only Left</div></div>
        <div class="stat"><div class="num green">{{ summary.unmatched_rm }}</div><div class="lbl">Unmatched RM</div></div>
      </div>
    {% endif %}

    <div class="actions">
      <a class="btn btn-primary" href="{{ url_for('download', rid=rid) }}">&#11015;&#65039; Download Excel Report</a>
      <a class="btn btn-outline" href="{{ url_for('dashboard', rid=rid) }}">&#128202; View Interactive Dashboard</a>
    </div>

    <a class="back" href="{{ url_for('index') }}">&larr; Generate another report</a>
  </div>
</body>
</html>
"""


# ----------------------------------------------------------------------
# Page 3: Dashboard (charts) — different layout per report type
# ----------------------------------------------------------------------

DASHBOARD_PAGE = """
<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Dashboard</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
""" + BASE_STYLE + """
  body { padding: 20px 24px 48px; }
  .topbar {
    max-width: 1280px; margin: 0 auto 16px; display: flex; align-items: center;
    justify-content: space-between; flex-wrap: wrap; gap: 14px;
  }
  .titles h1 { font-size: 1.25rem; margin: 0 0 3px; letter-spacing: -0.02em; }
  .titles .sub { color: var(--muted); font-size: 0.82rem; }
  .top-actions { display: flex; gap: 10px; }
  a.btn, button.btn {
    display: inline-flex; align-items: center; gap: 6px; padding: 10px 16px;
    border-radius: 9px; font-weight: 700; font-size: 0.85rem; text-decoration: none;
    border: none; cursor: pointer; transition: filter .15s ease;
  }
  .btn-primary { background: var(--accent); color: #04140d; }
  .btn-primary:hover { filter: brightness(1.08); }
  .btn-outline { background: transparent; color: var(--accent); border: 1.5px solid var(--accent-dim); }
  .btn-outline:hover { background: var(--accent-soft); }

  .info-banner.wrap { max-width: 1280px; margin: 0 auto 16px; }

  .stats {
    max-width: 1280px; margin: 0 auto 22px; display: grid;
    grid-template-columns: repeat(4, 1fr); gap: 12px;
  }
  .stat {
    background: var(--panel); border: 1px solid var(--panel-border);
    border-radius: 12px; padding: 16px; text-align: center;
  }
  .stat .num { font-size: 1.6rem; font-weight: 800; letter-spacing: -0.02em; }
  .stat .num.green { color: var(--accent); }
  .stat .num.red { color: var(--danger); }
  .stat .num.amber { color: var(--warn); }
  .stat .lbl { font-size: 0.72rem; color: var(--muted); margin-top: 3px; text-transform: uppercase; letter-spacing: 0.03em; }

  .grid {
    max-width: 1280px; margin: 0 auto; display: grid;
    grid-template-columns: repeat(2, 1fr); gap: 16px;
  }
  .grid .wide { grid-column: 1 / -1; }
  .panel {
    background: var(--panel); border: 1px solid var(--panel-border);
    border-radius: 14px; padding: 18px 20px 12px;
  }
  .panel h3 {
    margin: 0 0 12px; font-size: 0.92rem; font-weight: 700; color: var(--text);
    display: flex; align-items: center; gap: 8px;
  }
  .panel h3 .sq { width: 9px; height: 9px; border-radius: 3px; background: var(--accent); display: inline-block; }
  .chart-wrap { position: relative; height: 280px; }
  .chart-wrap.tall { height: 340px; }

  @media (max-width: 860px) {
    .grid { grid-template-columns: 1fr; }
    .stats { grid-template-columns: repeat(2, 1fr); }
  }
</style>
</head>
<body>
  <div class="topbar">
    <div class="titles">
      <h1>{{ summary.title }}</h1>
      <div class="sub">{{ row_count }} record(s) · generated {{ summary.generated_at }}</div>
    </div>
    <div class="top-actions">
      <a class="btn btn-outline" href="{{ url_for('results', rid=rid) }}">&larr; Back</a>
      <a class="btn btn-primary" href="{{ url_for('download', rid=rid) }}">&#11015;&#65039; Download Excel</a>
    </div>
  </div>

  {% if summary.report_type == "raw" %}
  <div class="stats">
    <div class="stat"><div class="num green">{{ summary.completed_count }}</div><div class="lbl">Completed</div></div>
    <div class="stat"><div class="num red">{{ summary.rejected_count }}</div><div class="lbl">Rejected</div></div>
    <div class="stat"><div class="num amber">{{ summary.pending_count }}</div><div class="lbl">Pending</div></div>
    <div class="stat"><div class="num green">{{ summary.avg_turnaround if summary.avg_turnaround is not none else "—" }}</div><div class="lbl">Avg Days To Complete</div></div>
  </div>

  <div class="grid">
    <div class="panel">
      <h3><span class="sq"></span>Overall Status Split</h3>
      <div class="chart-wrap"><canvas id="chartStatus"></canvas></div>
    </div>
    <div class="panel">
      <h3><span class="sq"></span>Daily Trend by Status</h3>
      <div class="chart-wrap"><canvas id="chartTrend"></canvas></div>
    </div>
    <div class="panel wide">
      <h3><span class="sq"></span>Pending Sub-Stage Breakdown ({{ summary.pending_count }} pending records)</h3>
      <div class="chart-wrap tall"><canvas id="chartStage"></canvas></div>
    </div>
    <div class="panel">
      <h3><span class="sq"></span>Rejected — By Request Type ({{ summary.rejected_count }} rejected records)</h3>
      <div class="chart-wrap tall"><canvas id="chartReject"></canvas></div>
    </div>
    <div class="panel">
      <h3><span class="sq"></span>Top Source Employees (all statuses)</h3>
      <div class="chart-wrap tall"><canvas id="chartSourceEmp"></canvas></div>
    </div>
    <div class="panel wide">
      <h3><span class="sq"></span>RM Wise (Top 12, by Status)</h3>
      <div class="chart-wrap tall"><canvas id="chartRMRaw"></canvas></div>
    </div>
    <div class="panel">
      <h3><span class="sq"></span>Account Mode</h3>
      <div class="chart-wrap"><canvas id="chartMode"></canvas></div>
    </div>
    <div class="panel wide">
      <h3><span class="sq"></span>Gender Split</h3>
      <div class="chart-wrap" style="max-width:340px;margin:0 auto;"><canvas id="chartGender"></canvas></div>
    </div>
  </div>

  {% else %}
  <div class="stats">
    <div class="stat"><div class="num green">{{ row_count }}</div><div class="lbl">Total Pending</div></div>
    <div class="stat"><div class="num green">{{ summary.esign_only }}</div><div class="lbl">Esign Only Left</div></div>
    <div class="stat"><div class="num green">{{ summary.unmatched_rm }}</div><div class="lbl">Unmatched RM</div></div>
    <div class="stat"><div class="num green">{{ summary.stage.labels|length }}</div><div class="lbl">Pending Stages</div></div>
  </div>

  <div class="grid">
    <div class="panel">
      <h3><span class="sq"></span>Pending By Stage</h3>
      <div class="chart-wrap"><canvas id="chartStage"></canvas></div>
    </div>
    <div class="panel">
      <h3><span class="sq"></span>Daily Creation Trend</h3>
      <div class="chart-wrap"><canvas id="chartTrend"></canvas></div>
    </div>
    <div class="panel">
      <h3><span class="sq"></span>Pending By L1 (Reporting Head)</h3>
      <div class="chart-wrap tall"><canvas id="chartL1"></canvas></div>
    </div>
    <div class="panel">
      <h3><span class="sq"></span>Pending By L2</h3>
      <div class="chart-wrap tall"><canvas id="chartL2"></canvas></div>
    </div>
    <div class="panel wide">
      <h3><span class="sq"></span>RM Wise (Top 15)</h3>
      <div class="chart-wrap tall"><canvas id="chartRM"></canvas></div>
    </div>
    <div class="panel">
      <h3><span class="sq"></span>Pending By Branch <span style="font-weight:400;color:var(--muted);">("Direct" = no RM, per Main.xlsx)</span></h3>
      <div class="chart-wrap tall"><canvas id="chartBranch"></canvas></div>
    </div>
    <div class="panel">
      <h3><span class="sq"></span>Account Mode</h3>
      <div class="chart-wrap"><canvas id="chartMode"></canvas></div>
    </div>
    <div class="panel wide">
      <h3><span class="sq"></span>Gender Split</h3>
      <div class="chart-wrap" style="max-width:340px;margin:0 auto;"><canvas id="chartGender"></canvas></div>
    </div>
  </div>
  {% endif %}

<script>
  const summary = {{ summary_json|safe }};

  // Varied multi-hue palette for category charts (L1/L2/Branch/RM/Mode/Gender/etc).
  // STATUS_COLORS stays separate — Completed/Rejected/Pending keep fixed
  // semantic colors wherever a chart is literally about those 3 statuses.
  const PALETTE = [
    "#3ddc97", "#5fb3ff", "#f4c95d", "#ff6b6b", "#b388ff",
    "#ff9f6b", "#4dd0e1", "#f472b6", "#a3e635", "#fb923c",
    "#818cf8", "#2dd4bf", "#facc15", "#f87171", "#c084fc",
  ];
  const STATUS_COLORS = { "Completed": "#3ddc97", "Rejected": "#ff6b6b", "Pending": "#f4c95d", "Unknown": "#8b98a5" };
  const GRID_COLOR = "rgba(230,237,243,0.06)";
  const TEXT_MUTED = "#8b98a5";

  Chart.defaults.color = TEXT_MUTED;
  Chart.defaults.font.family = "-apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Inter, sans-serif";
  Chart.defaults.borderColor = GRID_COLOR;

  function palette(n) {
    const out = [];
    for (let i = 0; i < n; i++) out.push(PALETTE[i % PALETTE.length]);
    return out;
  }
  function baseGrid() { return { grid: { color: GRID_COLOR }, ticks: { color: TEXT_MUTED } }; }

  new Chart(document.getElementById('chartStage'), {
    type: 'doughnut',
    data: {
      labels: summary.stage.labels,
      datasets: [{ data: summary.stage.values, backgroundColor: palette(summary.stage.labels.length), borderColor: '#0b0f14', borderWidth: 2 }]
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: { legend: { position: 'right', labels: { boxWidth: 12, padding: 10, font: { size: 11 } } } }
    }
  });

  if (summary.report_type === 'raw') {
    new Chart(document.getElementById('chartStatus'), {
      type: 'doughnut',
      data: {
        labels: summary.status.labels,
        datasets: [{ data: summary.status.values, backgroundColor: summary.status.labels.map(l => STATUS_COLORS[l] || '#8b98a5'), borderColor: '#0b0f14', borderWidth: 2 }]
      },
      options: {
        responsive: true, maintainAspectRatio: false,
        plugins: { legend: { position: 'right', labels: { boxWidth: 12, padding: 10, font: { size: 11 } } } }
      }
    });

    new Chart(document.getElementById('chartTrend'), {
      type: 'line',
      data: {
        labels: summary.trend.labels,
        datasets: ['Completed', 'Rejected', 'Pending'].map(s => ({
          label: s, data: summary.trend.series[s],
          borderColor: STATUS_COLORS[s], backgroundColor: STATUS_COLORS[s] + '26',
          fill: false, tension: 0.35, pointRadius: 1.5
        }))
      },
      options: {
        responsive: true, maintainAspectRatio: false,
        plugins: { legend: { position: 'bottom', labels: { boxWidth: 12, font: { size: 11 } } } },
        scales: { x: baseGrid(), y: baseGrid() }
      }
    });

    new Chart(document.getElementById('chartReject'), {
      type: 'bar',
      data: {
        labels: summary.reject_reasons.labels,
        datasets: [{ label: 'Rejected', data: summary.reject_reasons.values, backgroundColor: palette(summary.reject_reasons.labels.length), borderRadius: 5 }]
      },
      options: {
        indexAxis: 'y', responsive: true, maintainAspectRatio: false,
        plugins: { legend: { display: false } },
        scales: { x: baseGrid(), y: baseGrid() }
      }
    });

    new Chart(document.getElementById('chartSourceEmp'), {
      type: 'bar',
      data: {
        labels: summary.source_emp.labels,
        datasets: [{ label: 'Records', data: summary.source_emp.values, backgroundColor: palette(summary.source_emp.labels.length), borderRadius: 5 }]
      },
      options: {
        indexAxis: 'y', responsive: true, maintainAspectRatio: false,
        plugins: { legend: { display: false } },
        scales: { x: baseGrid(), y: baseGrid() }
      }
    });

    new Chart(document.getElementById('chartRMRaw'), {
      type: 'bar',
      data: {
        labels: summary.rm.labels,
        datasets: ['Completed', 'Rejected', 'Pending'].map(s => ({
          label: s, data: summary.rm.series[s], backgroundColor: STATUS_COLORS[s], borderRadius: 4
        }))
      },
      options: {
        responsive: true, maintainAspectRatio: false,
        plugins: { legend: { position: 'bottom', labels: { boxWidth: 12, font: { size: 11 } } } },
        scales: {
          x: { stacked: true, grid: { color: GRID_COLOR }, ticks: { color: TEXT_MUTED } },
          y: { stacked: true, grid: { color: GRID_COLOR }, ticks: { color: TEXT_MUTED } },
        }
      }
    });
  } else {
    new Chart(document.getElementById('chartTrend'), {
      type: 'line',
      data: {
        labels: summary.trend.labels,
        datasets: [{
          label: 'Records created', data: summary.trend.values,
          borderColor: '#3ddc97', backgroundColor: 'rgba(61,220,151,0.15)',
          fill: true, tension: 0.35, pointRadius: 2, pointBackgroundColor: '#3ddc97'
        }]
      },
      options: {
        responsive: true, maintainAspectRatio: false,
        plugins: { legend: { display: false } },
        scales: { x: baseGrid(), y: baseGrid() }
      }
    });

    new Chart(document.getElementById('chartL1'), {
      type: 'bar',
      data: {
        labels: summary.l1.labels,
        datasets: [{ label: 'Pending', data: summary.l1.values, backgroundColor: palette(summary.l1.labels.length), borderRadius: 5 }]
      },
      options: {
        indexAxis: 'y', responsive: true, maintainAspectRatio: false,
        plugins: { legend: { display: false } },
        scales: { x: baseGrid(), y: baseGrid() }
      }
    });

    new Chart(document.getElementById('chartL2'), {
      type: 'bar',
      data: {
        labels: summary.l2.labels,
        datasets: [{ label: 'Pending', data: summary.l2.values, backgroundColor: palette(summary.l2.labels.length), borderRadius: 5 }]
      },
      options: {
        indexAxis: 'y', responsive: true, maintainAspectRatio: false,
        plugins: { legend: { display: false } },
        scales: { x: baseGrid(), y: baseGrid() }
      }
    });

    new Chart(document.getElementById('chartRM'), {
      type: 'bar',
      data: {
        labels: summary.rm.labels,
        datasets: [{ label: 'Pending', data: summary.rm.values, backgroundColor: palette(summary.rm.labels.length), borderRadius: 5 }]
      },
      options: {
        responsive: true, maintainAspectRatio: false,
        plugins: { legend: { display: false } },
        scales: { x: baseGrid(), y: baseGrid() }
      }
    });

    new Chart(document.getElementById('chartBranch'), {
      type: 'bar',
      data: {
        labels: summary.branch.labels,
        datasets: [{ label: 'Pending', data: summary.branch.values, backgroundColor: palette(summary.branch.labels.length), borderRadius: 5 }]
      },
      options: {
        indexAxis: 'y', responsive: true, maintainAspectRatio: false,
        plugins: { legend: { display: false } },
        scales: { x: baseGrid(), y: baseGrid() }
      }
    });
  }

  new Chart(document.getElementById('chartMode'), {
    type: 'bar',
    data: {
      labels: summary.mode.labels,
      datasets: [{ label: 'Records', data: summary.mode.values, backgroundColor: palette(summary.mode.labels.length), borderRadius: 5 }]
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: { legend: { display: false } },
      scales: { x: baseGrid(), y: baseGrid() }
    }
  });

  new Chart(document.getElementById('chartGender'), {
    type: 'pie',
    data: {
      labels: summary.gender.labels,
      datasets: [{ data: summary.gender.values, backgroundColor: palette(summary.gender.labels.length), borderColor: '#0b0f14', borderWidth: 2 }]
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: { legend: { position: 'bottom', labels: { boxWidth: 12, font: { size: 11 } } } }
    }
  });
</script>
</body>
</html>
"""


# ----------------------------------------------------------------------
# Flask routes
# ----------------------------------------------------------------------

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        mode = request.form.get("mode", "pending")

        if mode == "raw":
            main_file = request.files.get("main_file_raw")
            if not main_file or main_file.filename == "":
                return render_template_string(PAGE, error="Please upload Main.xlsx.")
            l1l2l3_file = None
        else:
            main_file = request.files.get("main_file")
            l1l2l3_file = request.files.get("l1l2l3_file")
            if not main_file or main_file.filename == "" or not l1l2l3_file or l1l2l3_file.filename == "":
                return render_template_string(PAGE, error="Please upload both files.")

        try:
            main_bytes = io.BytesIO(main_file.read())
            if mode == "raw":
                output_buf, row_count, summary = generate_raw_report(main_bytes)
            else:
                l1l2l3_bytes = io.BytesIO(l1l2l3_file.read())
                output_buf, row_count, summary = generate_pending_report(main_bytes, l1l2l3_bytes)
        except Exception as exc:  # noqa: BLE001
            return render_template_string(
                PAGE, error=f"Failed to generate report: {exc}"
            )

        rid = uuid.uuid4().hex[:12]
        suffix = "Status_Report" if mode == "raw" else "Pending_Report"
        filename = f"ESIGN_{suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        REPORTS[rid] = {
            "bytes": output_buf.getvalue(),
            "filename": filename,
            "summary": summary,
            "row_count": row_count,
        }
        return redirect(url_for("results", rid=rid))
    return render_template_string(PAGE)


@app.route("/results/<rid>")
def results(rid):
    data = REPORTS.get(rid)
    if not data:
        return redirect(url_for("index"))
    return render_template_string(
        RESULTS_PAGE, rid=rid, summary=data["summary"], row_count=data["row_count"],
    )


@app.route("/dashboard/<rid>")
def dashboard(rid):
    data = REPORTS.get(rid)
    if not data:
        return redirect(url_for("index"))
    return render_template_string(
        DASHBOARD_PAGE,
        rid=rid,
        summary=data["summary"],
        row_count=data["row_count"],
        summary_json=json.dumps(data["summary"]),
    )


@app.route("/download/<rid>")
def download(rid):
    data = REPORTS.get(rid)
    if not data:
        return redirect(url_for("index"))
    buf = io.BytesIO(data["bytes"])
    return send_file(
        buf,
        as_attachment=True,
        download_name=data["filename"],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    # use_reloader=False avoids a crash on some Windows / Python 3.14 setups
    # where Flask's auto-reloader tries to register a signal handler outside
    # the main thread ("signal only works in main thread of the main
    # interpreter"). debug=True still gives you in-browser tracebacks.
    app.run(debug=True, use_reloader=False, host="0.0.0.0", port=5000)